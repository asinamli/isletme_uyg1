
"""ennnnn güncelll"""
import tkinter as tk
import openpyxl
import os

# Global değişkeni tanımlayın
if os.path.exists("counter.txt"):
    with open("counter.txt", "r") as f:
        a = int(f.read().strip())  # Dosyadan son değeri oku
else:
    a = 1  # İlk satır 1'den başlasın

def ana_form():
    global form
    form = tk.Tk()
    form.title('Ana Form')
    form.geometry('400x300')

    # Randevu Oluşturma Butonu
    btn_randevu = tk.Button(form, text='Randevu Oluştur', command=randevu_olustur)
    btn_randevu.pack(pady=10)

    # Müşteri Kaydetme Butonu
    btn_kaydet = tk.Button(form, text='Müşteri Kaydet', command=musteri_kaydet)
    btn_kaydet.pack(pady=10)

    # Müşteri Silme Butonu
    btn_sil = tk.Button(form, text='Müşteri Sil', command=musteri_sil)
    btn_sil.pack(pady=10)

    # Çıkış Butonu
    btn_cikis = tk.Button(form, text='Çıkış', command=form.quit)
    btn_cikis.pack(pady=10)

    form.mainloop()

def randevu_olustur():
    randevu_form = tk.Toplevel(form)
    randevu_form.title('Randevu Oluştur')
    randevu_form.geometry('400x400')

    # İsim için etiket ve giriş alanı
    tk.Label(randevu_form, text='Müşteri Adı').pack(pady=5)
    isim_entry = tk.Entry(randevu_form)
    isim_entry.pack(pady=5)

    tk.Label(randevu_form, text='Müşteri Soyadı').pack(pady=5)
    soyad_entry = tk.Entry(randevu_form)
    soyad_entry.pack(pady=5)

    tk.Label(randevu_form, text='Randevu Tarihi').pack(pady=5)
    tarih_entry = tk.Entry(randevu_form)
    tarih_entry.pack(pady=5)

    # Checkbox değişkenlerini tanımla
    geldi_var = tk.IntVar()
    gelmedi_var = tk.IntVar()

    tk.Checkbutton(randevu_form, text='Geldi', variable=geldi_var).pack()
    tk.Checkbutton(randevu_form, text='Gelmedi', variable=gelmedi_var).pack()

    # Randevu Ara Butonu
    def ara_randevu():
        isim = isim_entry.get()
        soyad = soyad_entry.get()

        # Excel dosyasını yükle
        dosya = openpyxl.load_workbook("./Deneme.xlsx")
        sayfa = dosya["Sheet1"]

        bulunan = False
        for row in range(1, a):
            if sayfa.cell(row=row, column=2).value == isim and sayfa.cell(row=row, column=1).value == soyad:
                bulunan = True
                tarih = sayfa.cell(row=row, column=3).value
                sonuc.config(text=f"BULUNAN MÜŞTERİ: {isim} {soyad}, Randevu Tarihi: {tarih}")
                break

        if not bulunan:
            sonuc.config(text="Müşteri bulunamadı.")

    tk.Button(randevu_form, text='Randevu Ara', command=ara_randevu).pack(pady=10)

    def kaydet_randevu():
        isim = isim_entry.get()
        soyad = soyad_entry.get()

        # Excel dosyasını yükle
        dosya = openpyxl.load_workbook("./Deneme.xlsx")
        sayfa = dosya["Sayfa1"]

        bulunan = False
        global a
        for row in range(1, a):
            if sayfa.cell(row=row, column=2).value == isim and sayfa.cell(row=row, column=1).value == soyad:
                bulunan = True
                tarih = sayfa.cell(row=row, column=3).value
                geldi = sayfa.cell(row=row, column=4).value

                sonuc.config(text=f"BULUNAN MÜŞTERİ: {isim} {soyad}, Randevu Tarihi: {tarih}, Durum: {geldi}")
                break

        if not bulunan:
            sonuc.config(text="Müşteri bulunamadı. Yeni randevu kaydediliyor.")

            # Yeni randevu kaydetme
            sayfa.cell(row=a, column=1, value=soyad)  # Soyad
            sayfa.cell(row=a, column=2, value=isim)    # Ad
            sayfa.cell(row=a, column=3, value=tarih_entry.get())   # Tarih
            if geldi_var.get():
                sayfa.cell(row=a, column=4, value="+")  # Gelirse
            elif gelmedi_var.get():
                sayfa.cell(row=a, column=4, value="-")  # Gelmezse
            a += 1  # Satırı bir artır

            dosya.save("./Deneme.xlsx")

            # Giriş alanlarını temizle
            isim_entry.delete(0, tk.END)
            soyad_entry.delete(0, tk.END)
            tarih_entry.delete(0, tk.END)
            geldi_var.set(0)
            gelmedi_var.set(0)

        # Güncellenen değeri dosyaya kaydet
        with open("counter.txt", "w") as f:
            f.write(str(a))

    tk.Button(randevu_form, text='Kaydet', command=kaydet_randevu).pack(pady=10)

    sonuc = tk.Label(randevu_form, text='', font='Times 12 bold', fg='red')
    sonuc.pack(pady=10)

def musteri_kaydet():
    global form
    form = tk.Toplevel(form)
    form.title('Müşteri Kaydet')
    form.geometry('600x450+650+350')
    form.minsize(450, 400)
    form.maxsize(550, 500)

    tk.Label(form, text='MÜŞTERİ KAYDET', font='Helvetica 15 bold').pack(pady=10)

    tk.Label(form, text='AD', font='Times 12').pack(pady=5)
    giris = tk.Entry(form, fg='black', font='Times 12 bold')
    giris.pack(pady=5)

    tk.Label(form, text='SOYAD', font='Times 12').pack(pady=5)
    giris2 = tk.Entry(form, fg='black', font='Times 12 bold')
    giris2.pack(pady=5)

    tk.Label(form, text='Tarih', font='Times 12').pack(pady=5)
    giris3 = tk.Entry(form, fg='black', font='Times 12 bold')
    giris3.pack(pady=5)

    # Checkbox değişkenlerini tanımla
    geldi_var = tk.IntVar()
    gelmedi_var = tk.IntVar()

    tk.Checkbutton(form, text='Geldi', variable=geldi_var).pack()
    tk.Checkbutton(form, text='Gelmedi', variable=gelmedi_var).pack()

    def kaydet():
        global a  # Global değişkeni kullan
        veri = giris.get()
        veri2 = giris2.get()
        veri3 = giris3.get()

        # Excel dosyasını yükle
        dosya = openpyxl.load_workbook("./Deneme.xlsx")
        sayfa = dosya["Sheet1"]

        # Satırı güncelle
        sayfa.cell(row=a, column=1, value=veri2)  # Soyad
        sayfa.cell(row=a, column=2, value=veri)    # Ad
        sayfa.cell(row=a, column=3, value=veri3)   # Tarih
        if geldi_var.get():
            sayfa.cell(row=a, column=4, value="+")  # Gelirse
        elif gelmedi_var.get():
            sayfa.cell(row=a, column=4, value="-")  # Gelmezse

        dosya.save("./Deneme.xlsx")

        # Satırı bir artır
        a += 1

        # Giriş alanlarını temizle
        giris.delete(0, tk.END)
        giris2.delete(0, tk.END)
        giris3.delete(0, tk.END)
        geldi_var.set(0)  # Checkbox'ları sıfırla
        gelmedi_var.set(0)

        # Güncellenen değeri dosyaya kaydet
        with open("counter.txt", "w") as f:
            f.write(str(a))

    tk.Button(form, text='Kaydet', command=kaydet).pack(pady=10)

def musteri_sil():
    global form
    form = tk.Toplevel(form)
    form.title('Müşteri Sil')
    form.geometry('400x300')

    tk.Label(form, text='MÜŞTERİ SİL', font='Helvetica 15 bold').pack(pady=10)

    tk.Label(form, text='AD', font='Times 12').pack(pady=5)
    giris = tk.Entry(form, fg='black', font='Times 12 bold')
    giris.pack(pady=5)

    tk.Label(form, text='SOYAD', font='Times 12').pack(pady=5)
    giris2 = tk.Entry(form, fg='black', font='Times 12 bold')
    giris2.pack(pady=5)

    def sil():
        global a  # Global değişkeni kullan
        isim = giris.get()
        soyad = giris2.get()

        # Excel dosyasını yükle
        dosya = openpyxl.load_workbook("./Deneme.xlsx")
        sayfa = dosya["Sheet1"]

        bulunan = False
        for row in range(1, a):  # 1'den a'ya kadar kontrol et
            if sayfa.cell(row=row, column=2).value == isim and sayfa.cell(row=row, column=1).value == soyad:
                bulunan = True
                # Satırı sil
                sayfa.delete_rows(row)
                a -= 1  # Satır sayısını bir azalt
                sonuc.config(text=f"{isim} {soyad} silindi.")  # Silindi mesajı
                break

        if not bulunan:
            sonuc.config(text="Müşteri bulunamadı")  # Müşteri bulunamazsa mesaj

        # Değişiklikleri kaydet
        dosya.save("./Deneme.xlsx")

    tk.Button(form, text='Sil', command=sil).pack(pady=10)

    sonuc = tk.Label(form, text='', font='Times 12 bold', fg='red')
    sonuc.pack(pady=10)

ana_form()
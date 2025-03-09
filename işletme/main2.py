import tkinter as tk
import openpyxl
import os

# Dosyada veri olup olmadığını kontrol etme ve a/b değişkenlerini güncelleme
def get_counter(file_path, sheet_name):
    if os.path.exists(file_path):
        dosya = openpyxl.load_workbook(file_path)
        sayfa = dosya[sheet_name]
        # Sayfada boş olmayan son satırı bul
        for row in range(sayfa.max_row, 0, -1):
            if any(sayfa.cell(row=row, column=col).value for col in range(1, sayfa.max_column + 1)):
                return row
        return 1  # Eğer sayfada hiç veri yoksa
    else:
        return 1  # Dosya yoksa 1. satırdan başla

# Sayfadaki boş satırları temizle
def temizle_boş_satırlar():
    dosya = openpyxl.load_workbook("./Deneme.xlsx")
    for sheet_name in ["Sayfa1", "Sayfa2"]:
        sayfa = dosya[sheet_name]
        for row in range(sayfa.max_row, 0, -1):
            if all(sayfa.cell(row=row, column=col).value is None for col in range(1, sayfa.max_column + 1)):
                sayfa.delete_rows(row)
    dosya.save("./Deneme.xlsx")

# Başlangıçta veri yazılacak satır numaralarını al
a = get_counter("./Deneme.xlsx", "Sayfa1")  # Randevu verilerini kaydetmek için
b = get_counter("./Deneme.xlsx", "Sayfa2")  # Müşteri verilerini kaydetmek için

# Ana formu global olarak tanımlıyoruz
form = None

def ana_form():
    global form
    form = tk.Tk()
    form.title('Ana Form')
    form.geometry('400x400')
    form.config(bg='#f0f0f0')

    header = tk.Frame(form, bg='#4CAF50')
    header.pack(fill=tk.X)

    title_label = tk.Label(header, text='Müşteri Yönetim Sistemi', bg='#4CAF50', fg='white', font=('Helvetica', 16, 'bold'))
    title_label.pack(pady=10)

    button_frame = tk.Frame(form, bg='#f0f0f0')
    button_frame.pack(pady=20)

    # Randevu Oluşturma Butonu
    btn_randevu = tk.Button(button_frame, text='Randevu Oluştur', command=randevu_olustur, bg='#4CAF50', fg='white', font=('Helvetica', 12))
    btn_randevu.pack(pady=10, padx=20, fill=tk.X)

    # Randevu Ara Butonu
    btn_ara = tk.Button(button_frame, text='Randevu Ara', command=randevu_ara, bg='#4CAF50', fg='white', font=('Helvetica', 12))
    btn_ara.pack(pady=10, padx=20, fill=tk.X)

    # Müşteri Kaydetme Butonu
    btn_kaydet = tk.Button(button_frame, text='Müşteri Kaydet', command=musteri_kaydet, bg='#4CAF50', fg='white', font=('Helvetica', 12))
    btn_kaydet.pack(pady=10, padx=20, fill=tk.X)

    # Müşteri Silme Butonu
    btn_sil = tk.Button(button_frame, text='Müşteri Sil', command=musteri_sil, bg='#4CAF50', fg='white', font=('Helvetica', 12))
    btn_sil.pack(pady=10, padx=20, fill=tk.X)

    # Çıkış Butonu
    btn_cikis = tk.Button(button_frame, text='Çıkış', command=form.quit, bg='#FF5722', fg='white', font=('Helvetica', 12))
    btn_cikis.pack(pady=10, padx=20, fill=tk.X)

    form.mainloop()

def randevu_olustur():

    randevu_form = tk.Toplevel(form)
    randevu_form.title('Randevu Oluştur')
    randevu_form.geometry('400x400')
    randevu_form.config(bg='#f0f0f0')

    tk.Label(randevu_form, text='Öğrenci Adı', bg='#f0f0f0', font=('Helvetica', 12)).pack(pady=5)
    isim_entry = tk.Entry(randevu_form)
    isim_entry.pack(pady=5, padx=20)

    tk.Label(randevu_form, text='Sınıf Bilgisi', bg='#f0f0f0', font=('Helvetica', 12)).pack(pady=5)
    soyad_entry = tk.Entry(randevu_form)
    soyad_entry.pack(pady=5, padx=20)

    tk.Label(randevu_form, text='Ders Tarihi', bg='#f0f0f0', font=('Helvetica', 12)).pack(pady=5)
    tarih_entry = tk.Entry(randevu_form)
    tarih_entry.pack(pady=5, padx=20)

    # Ücret alanı eklendi
    tk.Label(randevu_form, text='Ücret', bg='#f0f0f0', font=('Helvetica', 12)).pack(pady=5)
    ucret_entry = tk.Entry(randevu_form)
    ucret_entry.pack(pady=5, padx=20)

    sonuc = tk.Label(randevu_form, text='', font='Times 12 bold', fg='red', bg='#f0f0f0')
    sonuc.pack(pady=10)


    def kaydet_randevu():
        global a
        isim = isim_entry.get()
        soyad = soyad_entry.get()
        tarih = tarih_entry.get()
        ucret = ucret_entry.get()  # Ücret bilgisini alıyoruz

        # Müşteriyi Sayfa2'de kontrol et
        dosya = openpyxl.load_workbook("./Deneme.xlsx")
        sayfa_musteri = dosya["Sayfa2"]

        bulunan_musteri = False
        for row in range(1, b + 1):  # Sayfa2'deki satırlar üzerinden kontrol yap
            if sayfa_musteri.cell(row=row, column=1).value == soyad and sayfa_musteri.cell(row=row, column=2).value == isim:
                bulunan_musteri = True
                break

        if not bulunan_musteri:
            sonuc.config(text="Müşteri kayıtlı değil, randevu oluşturulamadı.")
            return  # Müşteri yoksa randevu oluşturulmaz

        # Randevu bilgilerini Sayfa1'e kaydet
        sayfa_randevu = dosya["Sayfa1"]
        # Boş bir satır bul ve veriyi oraya ekle
        for row in range(1, a + 2):  # Sayfa1'de boş bir satır arıyoruz
            if not any(sayfa_randevu.cell(row=row, column=col).value for col in range(1, 6)):
                sayfa_randevu.cell(row=row, column=1, value=isim)  # Ad
                sayfa_randevu.cell(row=row, column=2, value=soyad)  # Soyad
                sayfa_randevu.cell(row=row, column=3, value=tarih)  # Tarih
                sayfa_randevu.cell(row=row, column=4, value="gelmedi")  # Durum varsayılan "gelmedi"
                sayfa_randevu.cell(row=row, column=5, value=ucret)  # Ücret
                a = row  # Satır numarasını güncelle
                break

        # Dosyayı kaydet
        dosya.save("./Deneme.xlsx")

        # Giriş alanlarını temizle
        isim_entry.delete(0, tk.END)
        soyad_entry.delete(0, tk.END)
        tarih_entry.delete(0, tk.END)
        ucret_entry.delete(0, tk.END)  # Ücret alanını sıfırla

        # Güncellenen değeri dosyaya kaydet
        with open("counter.txt", "w") as f:
            f.write(str(a))

        sonuc.config(text="Randevu başarıyla oluşturuldu!")

    tk.Button(randevu_form, text='Kaydet', command=kaydet_randevu, bg='#4CAF50', fg='white', font=('Helvetica', 12)).pack(pady=10)

# Randevu Ara
def randevu_ara():
    ara_form = tk.Toplevel(form)
    ara_form.title('Randevu Ara')
    ara_form.geometry('500x500')
    ara_form.config(bg='#f0f0f0')

    tk.Label(ara_form, text='Öğrenci Adı', bg='#f0f0f0', font=('Helvetica', 12)).pack(pady=5)
    isim_entry = tk.Entry(ara_form)
    isim_entry.pack(pady=5, padx=20)

    tk.Label(ara_form, text='Sınıf Bilgisi', bg='#f0f0f0', font=('Helvetica', 12)).pack(pady=5)
    soyad_entry = tk.Entry(ara_form)
    soyad_entry.pack(pady=5, padx=20)

    sonuc = tk.Label(ara_form, text='', font='Times 12 bold', fg='red', bg='#f0f0f0')
    sonuc.pack(pady=10)

    def ara():
        isim = isim_entry.get().strip().lower()  # Boşlukları temizle ve küçük harfe dönüştür
        soyad = soyad_entry.get().strip().lower()  # Boşlukları temizle ve küçük harfe dönüştür

        dosya = openpyxl.load_workbook("./Deneme.xlsx")
        sayfa = dosya["Sayfa1"]

        bulunan = False
        for row in range(1, a + 1):  # Satırları doğru aralığa al
            soyad_cell = sayfa.cell(row=row, column=1).value
            isim_cell = sayfa.cell(row=row, column=2).value

            if soyad_cell and isim_cell:  # Boş satırları göz ardı et
                if soyad_cell.strip().lower() == soyad and isim_cell.strip().lower() == isim:  # Karşılaştırmayı küçük harf ile yap
                    bulunan = True
                    tarih = sayfa.cell(row=row, column=3).value
                    durum = sayfa.cell(row=row, column=4).value
                    ucret = sayfa.cell(row=row, column=5).value

                    sonuc.config(text=f"Öğrenci: {isim} {soyad}, Ders Tarihi: {tarih}, Durum: {durum}, Ücret: {ucret}")

                    break

        if not bulunan:
            sonuc.config(text="Öğrenci bulunamadı.")

    tk.Button(ara_form, text='Ara', command=ara, bg='#4CAF50', fg='white', font=('Helvetica', 12)).pack(pady=10)

# Müşteri Kaydet
def musteri_kaydet():
    kaydet_form = tk.Toplevel(form)
    kaydet_form.title('Müşteri Kaydet')
    kaydet_form.geometry('600x450')
    kaydet_form.config(bg='#f0f0f0')

    tk.Label(kaydet_form, text='MÜŞTERİ KAYDET', font=('Helvetica', 15, 'bold'), bg='#f0f0f0').pack(pady=10)

    tk.Label(kaydet_form, text='AD', font=('Times 12'), bg='#f0f0f0').pack(pady=5)
    giris = tk.Entry(kaydet_form, fg='black', font=('Times 12 bold'))
    giris.pack(pady=5, padx=20)

    tk.Label(kaydet_form, text='SOYAD', font=('Times 12'), bg='#f0f0f0').pack(pady=5)
    giris2 = tk.Entry(kaydet_form, fg='black', font=('Times 12 bold'))
    giris2.pack(pady=5, padx=20)

    tk.Label(kaydet_form, text='Tarih', font=('Times 12'), bg='#f0f0f0').pack(pady=5)
    giris3 = tk.Entry(kaydet_form, fg='black', font=('Times 12 bold'))
    giris3.pack(pady=5, padx=20)

    def kaydet():
        global b

        veri = giris.get()
        veri2 = giris2.get()
        veri3 = giris3.get()

        dosya = openpyxl.load_workbook("./Deneme.xlsx")
        sayfa = dosya["Sayfa2"]

        # Yeni veriyi son satıra ekleme
        for row in range(1, b + 2):  # Boş satır bul ve veriyi ekle
            if not any(sayfa.cell(row=row, column=col).value for col in range(1, sayfa.max_column + 1)):
                sayfa.cell(row=row, column=1, value=veri2)  # Soyad
                sayfa.cell(row=row, column=2, value=veri)   # Ad
                sayfa.cell(row=row, column=3, value=veri3)   # Tarih
                b = row
                break

        # Satır numarasını bir arttır
        dosya.save("./Deneme.xlsx")

        giris.delete(0, tk.END)
        giris2.delete(0, tk.END)
        giris3.delete(0, tk.END)

        # Güncellenen değeri dosyaya kaydet
        with open("counter2.txt", "w") as f:
            f.write(str(b))

    tk.Button(kaydet_form, text='Kaydet', command=kaydet, bg='#4CAF50', fg='white', font=('Helvetica', 12)).pack(pady=10)

# Müşteri Sil
def musteri_sil():
    sil_form = tk.Toplevel(form)
    sil_form.title('Müşteri Sil')
    sil_form.geometry('400x300')
    sil_form.config(bg='#f0f0f0')

    tk.Label(sil_form, text='MÜŞTERİ SİL', font=('Helvetica', 15, 'bold'), bg='#f0f0f0').pack(pady=10)

    tk.Label(sil_form, text='AD', font=('Times 12'), bg='#f0f0f0').pack(pady=5)
    giris = tk.Entry(sil_form, fg='black', font=('Times 12 bold'))
    giris.pack(pady=5, padx=20)

    tk.Label(sil_form, text='SOYAD', font=('Times 12'), bg='#f0f0f0').pack(pady=5)
    giris2 = tk.Entry(sil_form, fg='black', font=('Times 12 bold'))
    giris2.pack(pady=5, padx=20)

    sonuc = tk.Label(sil_form, text='', font='Times 12 bold', fg='red', bg='#f0f0f0')
    sonuc.pack(pady=10)

    def sil():
        global a
        isim = giris.get()
        soyad = giris2.get()

        dosya = openpyxl.load_workbook("./Deneme.xlsx")
        sayfa = dosya["Sayfa1"]

        bulunan = False
        for row in range(1, a):
            if sayfa.cell(row=row, column=2).value == isim and sayfa.cell(row=row, column=1).value == soyad:
                bulunan = True
                sayfa.delete_rows(row)
                a -= 1
                sonuc.config(text=f"{isim} {soyad} silindi.")
                break

        if not bulunan:
            sonuc.config(text="Müşteri bulunamadı")

        dosya.save("./Deneme.xlsx")

    tk.Button(sil_form, text='Sil', command=sil, bg='#FF5722', fg='white', font=('Helvetica', 12)).pack(pady=10)

# Program çalışırken boş satırları temizlemek için:
temizle_boş_satırlar()


ana_form()
